"""
Authors: Andreas Olborg and Jon Grendstad
Group: group_4
"""

from Game import Game
from Move import Move
import os
import re
import time
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
import textwrap

class Database:
    '''
    Encapsulate all games parsed from a PGN file or a Excel file.
    '''

    def __init__(self, games=None):
        if games is not None:
            self.games = games
        else: 
            self.games = []



    ## GETTERS ##

    def get_list_of_games(self):
        return self.games
    
    def get_list_of_white_wins(self):
        white_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1-0':
                white_wins.append(game)
        return white_wins
    
    def get_list_of_black_wins(self):
        black_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '0-1':
                black_wins.append(game)
        return black_wins
    
    def get_list_of_draws(self):
        draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2':
                draws.append(game)
        return draws
    
    def get_precentage_of_black_wins(self):
        return round((len(self.get_list_of_black_wins()) / len(self.get_list_of_games()))*100,2)
    
    def get_precentage_of_white_wins(self):
        return round((len(self.get_list_of_white_wins()) / len(self.get_list_of_games()))*100,2)
    
    def get_precentage_of_draws(self):
        return round((len(self.get_list_of_draws()) / len(self.get_list_of_games()))*100,2)

    def get_list_of_stockfish_wins_as_white(self):
        stockfish_wins = []
        for game in self.games:
            # i want it to lookup metadata where the value includes stockfish

            if game.lookup_meta_data('Result') == '1-0' and 'stockfish' in game.lookup_meta_data('White').lower():
                stockfish_wins.append(game)
        return stockfish_wins
    
    def get_list_of_stockfish_wins_as_black(self):
        stockfish_wins = []
        for game in self.games:
    
            if game.lookup_meta_data('Result') == '0-1' and 'stockfish' in game.lookup_meta_data('Black').lower():
                stockfish_wins.append(game)
        return stockfish_wins
    
    def get_list_of_stockfish_losses_as_white(self):
        stockfish_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '0-1' and 'stockfish' in game.lookup_meta_data('White').lower():
                stockfish_wins.append(game)
        return stockfish_wins
    
    def get_list_of_stockfish_losses_as_black(self):
        stockfish_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1-0' and 'stockfish' in game.lookup_meta_data('Black').lower():
                stockfish_wins.append(game)
        return stockfish_wins
    
    def get_list_of_stockfish_draws_as_white(self):
        stockfish_draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2' and 'stockfish' in game.lookup_meta_data('White').lower():
                stockfish_draws.append(game)
        return stockfish_draws
    
    def get_list_of_stockfish_draws_as_black(self):
        stockfish_draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2' and 'stockfish' in game.lookup_meta_data('Black').lower():
                stockfish_draws.append(game)
        return stockfish_draws
    
    def get_list_of_stockfish_wins(self):
        return self.get_list_of_stockfish_wins_as_black() + self.get_list_of_stockfish_wins_as_white()  
    
    def get_list_of_stockfish_losses(self):
        return self.get_list_of_stockfish_losses_as_black() + self.get_list_of_stockfish_losses_as_white()
    
    def get_list_of_games_where_stockfish_wins_or_draws(self):
        return self.get_list_of_stockfish_wins() + self.get_list_of_draws()
    
    def get_list_of_games_where_stockfish_is_white(self):
        stockfish_white = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('White'): 
                stockfish_white.append(game)
        return stockfish_white
    
    def get_list_of_games_where_stockfish_is_black(self):
        stockfish_black = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('Black'):
                stockfish_black.append(game)
        return stockfish_black

    def get_games_with_move_sequence(self, move_sequence): #move_sequence is a list of moves (e.g. ['e4', 'e5', 'Nf3', 'Nc6'])
        games_with_move_sequence = []
        for game in self.games:
            moves = game.get_moves_without_comments()
            if len(moves) >= len(move_sequence):
                for i in range(len(moves) - len(move_sequence) + 1):
                    if moves[i:i+len(move_sequence)] == move_sequence:
                        games_with_move_sequence.append(game)
                        break
        return games_with_move_sequence
    
    def get_list_with_opening(self, opening):
        games_with_opening = []
        for game in self.games:
            if game.lookup_meta_data('Opening') == opening:

                games_with_opening.append(game)

        return games_with_opening

    def get_opening_counts(self): #Returns a dictionary with keys being openings and values being the number of games with that opening
        openings = {}
        for game in self.games:
            opening = game.lookup_meta_data('Opening')
            if opening in openings:
                openings[opening] += 1      #If the opening is already in the dictionary, increment its value by 1
            else:
                openings[opening] = 1       #If the opening is not in the dictionary, add it with a value of 1
        return openings
    
    def get_openings_that_occurred_at_least_n_times(self, n):
        openings = self.get_opening_counts()                                    #Get a dictionary with keys being openings and values being the number of games with that opening
        openings_that_occurred_at_least_n_times = {}                                    #Create a new dictionary to store the openings that occurred at least n times
        for opening in openings:                            
            if openings[opening] >= n:                                                  #If the opening occurred at least n times, add it to the new dictionary
                openings_that_occurred_at_least_n_times[opening] = openings[opening]    #The value of the new dictionary is the number of games with that opening
        
        return openings_that_occurred_at_least_n_times

    def get_standard_deviation_of_moves(self):
        amount_of_moves = []
        for game in self.games:
            amount_of_moves.append(len(game.get_moves()))
        return np.std(amount_of_moves)

    def get_mean_number_of_moves(self):
        amount_of_moves = []
        for game in self.games:
            amount_of_moves.append(len(game.get_moves()))
        return np.mean(amount_of_moves)



    ## PARSE AND COMPOSE FOR PGN AND EXCEL ##

    def compose_to_excel(self, path, sheet_name):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        i = 1
        for game in self.games:

            for key, value in game.get_meta_data().items():
                worksheet.cell(row=i, column=1, value=key)
                worksheet.cell(row=i, column=2, value=value)
                i += 1
            i += 1  

            for move in game.get_moves():
                worksheet.cell(row=i, column=1, value=move.get_number())
                worksheet.cell(row=i, column=2, value=move.get_white_move())
                worksheet.cell(row=i, column=3, value=move.get_white_comment())
                worksheet.cell(row=i, column=4, value=move.get_black_move())
                worksheet.cell(row=i, column=5, value=move.get_black_comment())
                i += 1

            i += 1
        workbook.save(filename=path)

    def parse_from_excel(self, path, sheet_name):
        workbook = load_workbook(filename=path)
        worksheet = workbook[sheet_name]
        phase = 0
        
        meta_data = {}
        moves = []
        for row in worksheet.iter_rows(min_row=1, values_only=True):

            if phase == 0:
                if row[0] is None:
                    phase = 1
                    continue

                meta_data[row[0]] = row[1]

            if phase == 1:
                if row[0] is None:
                    phase = 2
                    continue

                number = row[0]
                white_move = row[1]
                white_move_comment = row[2]
                black_move = row[3]
                black_move_comment = row[4]

                moves.append(Move(number, white_move, white_move_comment, black_move, black_move_comment))

            if phase == 2:
                self.games.append(Game(meta_data, moves))
                meta_data = {}
                moves = []
                phase = 0
                continue

    def compose_to_pgn(self , path):
        pgn_data = ""
        for game in self.games:
            meta_data = ""
            for key, value in game.get_meta_data().items():
                line = "[" + str(key) + " \"" + str(value) + "\"]"
                meta_data = meta_data + line + "\n"
            
            pgn_data += meta_data + "\n"

            moves = ""
            for move in game.get_moves(): 
                moves += move.get_number() + ". "
                
                if move.get_white_move() is not None:
                    moves += move.get_white_move() + " "
               
                if move.get_white_comment() is not None:
                    moves += move.get_white_comment() + " "
               
                if move.get_black_move() is not None:
                    moves += move.get_black_move() + " "
             
                if move.get_black_comment() is not None:
                    moves += move.get_black_comment() + " " 
            
            moves = textwrap.fill(moves, width=80)
            score = game.get_meta_data()["Result"]
            pgn_data += moves + score + "\n\n"
        
        pgn_file = open(path,"w")
        pgn_file.write(pgn_data)
        pgn_file.close()
    
    def parse_from_pgn(self, path):
        file = open(path, 'r')
        pgn_data = file.read()
        games = list(filter(lambda x: len(x) > 0, pgn_data.split('\n\n[')))
        
        game_list = []
        
        for game in games:
            chessgame = Game({}, [])            
            sections = list(filter(lambda x: len(x) > 0, game.split('\n\n'))) # Split on empty lines
            meta_data = sections[0]
            meta_data = meta_data.split('\n')
            meta_data = list(filter(lambda x: len(x) > 0, meta_data))                       # Remove empty strings
            # Remove quatation marks
            meta_data = list(map(lambda x: x.replace('[', '').replace(']', ''), meta_data)) # Remove brackets
            meta_data = list(map(lambda x: x.split(' ', 1), meta_data))                     # Split on first space
            meta_data = list(map(lambda x: (x[0], x[1].replace('"', '')), meta_data))       # Remove quotes
            meta_data = dict(meta_data)
            for key, value in meta_data.items():
                chessgame.add_meta_data(key, value)
            
            moves = sections[1].replace('\n', ' ')
            
            # remove score at the end of the moves
            moves = re.sub(r'\s(1\/2-1\/2|1-0|0-1)$', '', moves)
            
            pattern = r'\d+\.\s[\S\s]+?(?=\d+\.\s|\Z)' # Matches all moves
            # Define a regular expression to match the input string
            matches = re.findall(pattern, moves)
            for match in matches:
                pattern = r'(\d+)\.?\s*(\S+)\s*({.*?})?\s*(\S+)?\s*({.*?})?'
                result = re.match(pattern, match)
                if result:
                    number = result.group(1)
                    white_move = result.group(2)
                    white_move_comment = result.group(3)
                    black_move = result.group(4)
                    black_move_comment = result.group(5)
                    chessgame.add_move(Move(number, white_move, white_move_comment, black_move, black_move_comment)) # Add move to game
                else:
                    print("No match---------------------")
            game_list.append(chessgame)

        self.games = game_list

def main():
    time_start = time.time()

    database = Database()

    # Parsing from pgn file
    database.parse_from_pgn("databases/Stockfish_15_64-bit.commented.[2600].pgn")

    # Composing from Database instance to pgn and excel format
    database.compose_to_pgn("databases/Stockfish_15_64-bit.commented.[2600]_composed.pgn")
    database.compose_to_excel("databases/Stockfish_15_64-bit.commented.[2600]_composed.xlsx", "Sheet1")

    # Paring from excel file
    database.parse_from_excel("databases/Stockfish_15_64-bit.commented.[2600]_composed.xlsx", "Sheet1")
    
    print(f"Time: {time.time() - time_start}")
    
if __name__ == "__main__":
    main()
    

    