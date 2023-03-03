from PGNGame import *
from PGNMove import *
import re
import time

import os
from docx import Document
from docx.shared import Inches

import numpy as np
import matplotlib.pyplot as plt

from openpyxl import Workbook
from openpyxl import load_workbook

class PGNDatabase:
    '''
    Encapsulate all games parsed from a PGN file.
    '''
    def __init__(self, path):
        self.games = self.parse(path)  
        self.path = path
        self.statistics = self.get_statistics()

    # def get_games(self):
    #     return self.games
    
    def get_statistics(self):
        '''
        Get statistics of the database.
        '''
        amount_of_stockfish_wins = 0
        amount_of_stockfish_loses = 0
        amount_of_stockfish_draws = 0
        amount_of_stockfish_is_white_and_wins = 0
        amount_of_stockfish_is_black_and_wins = 0
        amount_of_stockfish_is_white_and_loses = 0
        amount_of_stockfish_is_black_and_loses = 0
        amount_of_stockfish_is_white_and_draws = 0
        amount_of_stockfish_is_black_and_draws = 0

        all_games = []
        stockfish_white = []
        stockfish_black = []
        stockfish_draws = []
        stockfish_wins = []
        stockfish_loses = []
        stockfish_is_white_and_wins = []
        stockfish_is_black_and_wins = []
        stockfish_is_white_and_loses = []
        stockfish_is_black_and_loses = []
        stockfish_is_white_and_draws = []
        stockfish_is_black_and_draws = []
        
        for game in self.games:
            all_games.append(game)
            if "Stockfish" in game.lookup_meta_data('White'):
                stockfish_white.append(game)
                if game.lookup_meta_data('Result') == '1-0':
                    amount_of_stockfish_wins += 1
                    amount_of_stockfish_is_white_and_wins += 1
                    stockfish_wins.append(game)
                    stockfish_is_white_and_wins.append(game)
                elif game.lookup_meta_data('Result') == '0-1':
                    amount_of_stockfish_loses += 1
                    amount_of_stockfish_is_white_and_loses += 1
                    stockfish_loses.append(game)
                    stockfish_is_white_and_loses.append(game)
                elif game.lookup_meta_data('Result') == '1/2-1/2':
                    amount_of_stockfish_draws += 1
                    amount_of_stockfish_is_white_and_draws += 1
                    stockfish_draws.append(game)
                    stockfish_is_white_and_draws.append(game)
            elif "Stockfish" in game.lookup_meta_data('Black'):
                stockfish_black.append(game)
                if game.lookup_meta_data('Result') == '1-0':
                    amount_of_stockfish_wins += 1
                    amount_of_stockfish_is_black_and_wins += 1
                    stockfish_wins.append(game)
                    stockfish_is_black_and_wins.append(game)
                elif game.lookup_meta_data('Result') == '0-1':
                    amount_of_stockfish_loses += 1
                    amount_of_stockfish_is_black_and_loses += 1
                    stockfish_loses.append(game)
                    stockfish_is_black_and_loses.append(game)
                elif game.lookup_meta_data('Result') == '1/2-1/2':
                    amount_of_stockfish_draws += 1
                    amount_of_stockfish_is_black_and_draws += 1
                    stockfish_draws.append(game)
                    stockfish_is_black_and_draws.append(game)

        statistics = {
            "all_games": all_games,
            "stockfish_white": stockfish_white,
            "stockfish_black": stockfish_black,
            "stockfish_draws": stockfish_draws,
            "stockfish_wins": stockfish_wins,
            "stockfish_loses": stockfish_loses,
            "stockfish_is_white_and_wins": stockfish_is_white_and_wins,
            "stockfish_is_black_and_wins": stockfish_is_black_and_wins,
            "stockfish_is_white_and_loses": stockfish_is_white_and_loses,
            "stockfish_is_black_and_loses": stockfish_is_black_and_loses,
            "stockfish_is_white_and_draws": stockfish_is_white_and_draws,
            "stockfish_is_black_and_draws": stockfish_is_black_and_draws,
            "amount_of_stockfish_wins": amount_of_stockfish_wins,
            "amount_of_stockfish_loses": amount_of_stockfish_loses,
            "amount_of_stockfish_draws": amount_of_stockfish_draws,
            "amount_of_stockfish_is_white_and_wins": amount_of_stockfish_is_white_and_wins,
            "amount_of_stockfish_is_black_and_wins": amount_of_stockfish_is_black_and_wins,
            "amount_of_stockfish_is_white_and_loses": amount_of_stockfish_is_white_and_loses,
            "amount_of_stockfish_is_black_and_loses": amount_of_stockfish_is_black_and_loses,
            "amount_of_stockfish_is_white_and_draws": amount_of_stockfish_is_white_and_draws,
            "amount_of_stockfish_is_black_and_draws": amount_of_stockfish_is_black_and_draws,
        }
        return statistics
            



    # def get_white_wins(self):
    #     white_wins = []
    #     for game in self.games:
    #         if game.lookup_meta_data('Result') == '1-0':
    #             white_wins.append(game)
    #     return white_wins
    
    # def get_black_wins(self):
    #     black_wins = []
    #     for game in self.games:
    #         if game.lookup_meta_data('Result') == '0-1':
    #             black_wins.append(game)
    #     return black_wins
    
    # def get_draws(self):
    #     draws = []
    #     for game in self.games:
    #         if game.lookup_meta_data('Result') == '1/2-1/2':
    #             draws.append(game)
    #     return draws
    

    
    def get_games_with_move_sequence(self, move_sequence): #move_sequence is a list of moves (e.g. ['e4', 'e5', 'Nf3', 'Nc6'])
        games_with_move_sequence = []
        for game in self.games:
            moves = game.get_moves_without_comments()
            if len(moves) >= len(move_sequence):
                for i in range(len(moves) - len(move_sequence) + 1):
                    if moves[i:i+len(move_sequence)] == move_sequence:
                        games_with_move_sequence.append(game)
                        break
        return games_with_move_sequence
    
    
    # Checks on "OPENING" metadata
    def get_games_with_opening(self, opening):
        games_with_opening = []
        for game in self.games:
            if game.lookup_meta_data('Opening') == opening:
                games_with_opening.append(game)
        return games_with_opening
    
    def get_statistics_on_openings(self, list_of_games): #Returns a dictionary with keys being openings and values being the number of games with that opening
        openings = {}
        for game in list_of_games:
            opening = game.lookup_meta_data('Opening')
            if opening in openings:
                openings[opening] += 1      #If the opening is already in the dictionary, increment its value by 1
            else:
                openings[opening] = 1       #If the opening is not in the dictionary, add it with a value of 1
        return openings
    
    def get_openings_that_occurred_at_least_n_times(self, list_of_games, n):
        openings = self.get_statistics_on_openings(list_of_games)                                    #Get a dictionary with keys being openings and values being the number of games with that opening
        openings_that_occurred_at_least_n_times = {}                                    #Create a new dictionary to store the openings that occurred at least n times
        for opening in openings:                            
            if openings[opening] >= n:                                                  #If the opening occurred at least n times, add it to the new dictionary
                openings_that_occurred_at_least_n_times[opening] = openings[opening]    #The value of the new dictionary is the number of games with that opening
        return openings_that_occurred_at_least_n_times
        

    def get_eco_that_occurred_at_least_n_times(self, n):
        openings = self.get_statistics_on_eco()                                    #Get a dictionary with keys being openings and values being the number of games with that opening
        openings_that_occurred_at_least_n_times = {}                                    #Create a new dictionary to store the openings that occurred at least n times
        for opening in openings:                            
            if openings[opening] >= n:                                                  #If the opening occurred at least n times, add it to the new dictionary
                openings_that_occurred_at_least_n_times[opening] = openings[opening]    #The value of the new dictionary is the number of games with that opening
        return openings_that_occurred_at_least_n_times
    
    
    # Checks on "ECO" metadata
    def get_games_with_eco(self, opening):
        games_with_opening = []
        for game in self.games:
            if game.lookup_meta_data('ECO') == opening:
                games_with_opening.append(game)
        return games_with_opening
    
    def get_statistics_on_eco(self): #Returns a dictionary with keys being openings and values being the number of games with that opening
        openings = {}
        for game in self.games:
            opening = game.lookup_meta_data('ECO')
            if opening in openings:
                openings[opening] += 1      #If the opening is already in the dictionary, increment its value by 1
            else:
                openings[opening] = 1       #If the opening is not in the dictionary, add it with a value of 1
        return openings
    
    def get_games_where_stockfish_is_white(self):
        stockfish_white = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('White'): #Uses regex to check if "Stockfish" is in the string, so we can test for Stockfish 15 64-bit, Stockfish 15 32-bit, etc.
                stockfish_white.append(game)
        return stockfish_white
    
    def get_games_where_stockfish_is_black(self):
        stockfish_black = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('Black'):
                stockfish_black.append(game)
        return stockfish_black        
    


    # def get_stockfish_draws(self, list_of_drawed_games):
    #     stockfish_draws_as_white = [game for game in list_of_drawed_games if "Stockfish" in game.lookup_meta_data('White')]
    #     stockfish_draws_as_black = [game for game in list_of_drawed_games if "Stockfish" in game.lookup_meta_data('Black')]
    #     return stockfish_draws_as_white, stockfish_draws_as_black

    # #IF game.lookup_meta_data('White') includes "Stockfish"
    # def get_stockfish_wins(self, list_of_games):
    #     stockfish_wins_as_white = [game for game in list_of_games if "Stockfish" in game.lookup_meta_data('White') and game.lookup_meta_data('Result') == '1-0']
    #     stockfish_wins_as_black = [game for game in list_of_games if "Stockfish" in game.lookup_meta_data('Black') and game.lookup_meta_data('Result') == '0-1']
    #     return stockfish_wins_as_white, stockfish_wins_as_black

    # def get_stockfish_losses(self, list_of_games):
    #     stockfish_losses_as_white = [game for game in list_of_games if "Stockfish" in game.lookup_meta_data('White') and game.lookup_meta_data('Result') == '0-1']
    #     stockfish_losses_as_black = [game for game in list_of_games if "Stockfish" in game.lookup_meta_data('Black') and game.lookup_meta_data('Result') == '1-0']
    #     return stockfish_losses_as_white, stockfish_losses_as_black



    ##### TASK 8 ########
    def get_standard_deviation_of_moves(self, list_of_games):
        amount_of_moves = []
        for game in list_of_games:
            amount_of_moves.append(len(game.get_moves()))
        return np.std(amount_of_moves)

    def get_mean_number_of_moves(self, list_of_games):
        amount_of_moves = []
        for game in list_of_games:
            amount_of_moves.append(len(game.get_moves()))
        return np.mean(amount_of_moves)
    
    def get_move_count_distribution(self, list_of_games):
        move_count_distribution = {}
        for game in list_of_games:
            move_count = len(game.get_moves())
            if move_count in move_count_distribution:
                move_count_distribution[move_count] += 1
            else:
                move_count_distribution[move_count] = 1
        return move_count_distribution

    def get_plycount_distribution(self, list_of_games):
        plycount_distribution = {}
        for game in list_of_games:
            plycount = int(game.lookup_meta_data('PlyCount'))
            if plycount in plycount_distribution:
                plycount_distribution[plycount] += 1
            else:
                plycount_distribution[plycount] = 1
        return plycount_distribution

    def sort_dict(self, dict):
        return sorted(dict.items(), key=lambda x: x[0])
    

    def plot_plycount_distribution(self, list_of_games):
        plycount_distribution = self.get_move_count_distribution(list_of_games)
        plycount_distribution = self.sort_dict(plycount_distribution)
        #Clear plot
        x = []
        y = []
        for key, value in plycount_distribution:
            x.append(key)
            y.append(value)
        plt.figure(figsize=(10,3))
        plt.plot(x, y)
        plt.xlim(15, 250)
        plt.ylim(0, 150)
        plt.savefig('./plots/plycount_distribution.png')
        # plt.show()

    def plot_move_count_distribution(self, list_of_games):
        move_count_distribution = self.get_move_count_distribution(list_of_games)
        move_count_distribution = self.sort_dict(move_count_distribution)
        #Clear plot
        x = []
        y = []
        for key, value in move_count_distribution:
            x.append(key)
            y.append(value)
        #Min and max values for x and y axis
        plt.ylim(0, 150)        
        plt.xlim(15, 125)
        plt.figure(figsize=(50,100))
        plt.xlabel('Number of moves')
        plt.ylabel('Number of games')
        plt.plot(x, y)
        plt.fill_between(x, y, color='blue', alpha=0.5)
        plt.savefig('./plots/move_count_distribution.png')


    def plot_move_count_histogram_cumulative(self, list_of_games, textinfo, axis=None):
        move_count_distribution = self.get_move_count_distribution(list_of_games)
        move_count_distribution = self.sort_dict(move_count_distribution)
        data = [(moves, games) for moves, games in move_count_distribution]
        data = sorted([(moves, games) for moves, games in data], reverse=True)
        # calculate the cumulative sum of the games for each move count
        cumulative_data = []
        cumulative_sum = 0
        for moves, games in data:
            cumulative_sum += games
            cumulative_data.append((moves, cumulative_sum))
        # plot the cumulative frequency as a function of the move count
        x = [moves for moves, games in cumulative_data]
        y = [games for moves, games in cumulative_data]
        if axis is None:  # if no axis is given, create a new one
            axis = plt.gca() # get current axis
        axis.plot(x, y, label=textinfo)
        axis.set_xlim(15, 125)
        axis.set_xlabel('Number of Moves')
        axis.set_ylabel('Cumulative Number of Games')
        axis.set_title('Reverse Histogram of Chess Game Length')

    def plot_multiple_move_count_histogram_cumulative(self, dict_of_lists_of_games):
        for key, value in dict_of_lists_of_games.items():
            self.plot_move_count_histogram_cumulative(value, key)
        plt.legend()

    def save_histogram(self, path):
        plt.savefig(path)


    def clear_plot(self):
        plt.clf()

    def compose_to_excel(self):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = 'pgn_as_excel'

        i = 1
        for game in self.games:

            for key, value in game.get_meta_data().items():
                worksheet.cell(row=i, column=1, value=key)
                worksheet.cell(row=i, column=2, value=value)
                i += 1
            i += 1  

            for move in game.get_moves():
                worksheet.cell(row=i, column=1, value=move.get_number())
                worksheet.cell(row=i, column=2, value=move.get_white_move())
                worksheet.cell(row=i, column=3, value=move.get_white_comment())
                worksheet.cell(row=i, column=4, value=move.get_black_move())
                worksheet.cell(row=i, column=5, value=move.get_black_comment())
                i += 1

            i += 1
        workbook.save(filename='pgn_as_excel.xlsx')

    def parse_from_excel(self):
        workbook = load_workbook(filename='pgn_as_excel.xlsx')
        worksheet = workbook['pgn_as_excel']
        new_game=False
        phase = 0
        
        meta_data = {}
        moves = []
        for row in worksheet.iter_rows(min_row=1, values_only=True):

            if phase == 0:
                if row[0] is None:
                    phase = 1
                    continue

                meta_data[row[0]] = row[1]

            if phase == 1:
                if row[0] is None:
                    phase = 2
                    continue

                number = row[0]
                white_move = row[1]
                white_move_comment = row[2]
                black_move = row[3]
                black_move_comment = row[4]

                moves.append(PGNMove(number, white_move, white_move_comment, black_move, black_move_comment))

            if phase == 2:
                self.games.append(PGNGame(meta_data, moves))
                meta_data = {}
                moves = []
                phase = 0
                continue

    def compose(self):
        # TODO fix correct linebreak in the moves
        
        pgn_data = ""
        for game in self.games:
            meta_data = ""
            for key, value in game.get_meta_data().items():
                line = "[" + str(key) + " \"" + str(value) + "\"]"
                meta_data = meta_data + line + "\n"
            
            pgn_data += meta_data + "\n"

            moves = ""
            for move in game.get_moves(): 
                moves += move.get_number() + ". "
                
                if move.get_white_move() is not None:
                    moves += move.get_white_move() + " "
               
                if move.get_white_comment() is not None:
                    moves += move.get_white_comment() + " "
               
                if move.get_black_move() is not None:
                    moves += move.get_black_move() + " "
             
                if move.get_black_comment() is not None:
                    moves += move.get_black_comment() + " "
            
            score = game.get_meta_data()["Result"]
            pgn_data += moves + score + "\n\n"
        
        pgn_file = open("./databases/test.pgn","w")
        pgn_file.write(pgn_data)
        pgn_file.close()
    
    def parse(self,path):
        file = open(path, 'r')
        pgn_data = file.read()
        games = list(filter(lambda x: len(x) > 0, pgn_data.split('\n\n[')))
        
        game_list = []
        
        for game in games:
            chessgame = PGNGame({}, [])            
            sections = list(filter(lambda x: len(x) > 0, game.split('\n\n'))) # Split on empty lines
            meta_data = sections[0]
            meta_data = meta_data.split('\n')
            meta_data = list(filter(lambda x: len(x) > 0, meta_data))                       # Remove empty strings
            # Remove quatation marks
            meta_data = list(map(lambda x: x.replace('[', '').replace(']', ''), meta_data)) # Remove brackets
            meta_data = list(map(lambda x: x.split(' ', 1), meta_data))                     # Split on first space
            meta_data = list(map(lambda x: (x[0], x[1].replace('"', '')), meta_data))       # Remove quotes
            meta_data = dict(meta_data)
            for key, value in meta_data.items():
                chessgame.add_meta_data(key, value)
            
            moves = sections[1].replace('\n', ' ')
            
            # remove score at the end of the moves
            moves = re.sub(r'\s(1\/2-1\/2|1-0|0-1)$', '', moves)
            
            pattern = r'\d+\.\s[\S\s]+?(?=\d+\.\s|\Z)' # Matches all moves
            # Define a regular expression to match the input string
            matches = re.findall(pattern, moves)
            for match in matches:
                pattern = r'(\d+)\.?\s*(\S+)\s*({.*?})?\s*(\S+)?\s*({.*?})?'
                result = re.match(pattern, match)
                if result:
                    number = result.group(1)
                    white_move = result.group(2)
                    white_move_comment = result.group(3)
                    black_move = result.group(4)
                    black_move_comment = result.group(5)
                    chessgame.add_move(PGNMove(number, white_move, white_move_comment, black_move, black_move_comment)) # Add move to game
                else:
                    print("No match---------------------")
            game_list.append(chessgame)

        return game_list
            
def test():
    start_time = time.time()
    pgn = PGNDatabase("./databases/sample.pgn")
    print(pgn.games)
    for i, g in enumerate(pgn.get_games()):
        pass
        print("-----------------" + str(i) + "-----------------")
        print(g.meta_data)
        print("")
        for m in g.moves:
            pass
            print(m)
    
    
    pgn.games = []
    
    print(pgn.games)
    

    
    #pgn.compose_to_excel()
    pgn.parse_from_excel()

    for i, g in enumerate(pgn.get_games()):
        pass
        print("-----------------" + str(i) + "-----------------")
        print(g.meta_data)
        print("")
        for m in g.moves:
            pass
            print(m)
    
    print(f"Time: {time.time() - start_time}")

#test()
            
            

def main():
    start_time = time.time()
    # pgn = PGNDatabase("./databases/sample.pgn")
    pgn = PGNDatabase("./databases/Stockfish_15_64-bit.commented.[2600].pgn")

    all_games = pgn.statistics["all_games"]
    games_where_stockfish_is_white = pgn.statistics["stockfish_white"]
    games_where_stockfish_is_black = pgn.statistics["stockfish_black"]
    list_of_games = all_games
    
    one = pgn.get_openings_that_occurred_at_least_n_times(all_games, 1)
    # print(one)

    
    fig, ax = plt.subplots()
    fig.set_size_inches(10,5)
    
    pgn.plot_move_count_histogram_cumulative(all_games, "All games", axis=ax)
    pgn.plot_move_count_histogram_cumulative(games_where_stockfish_is_white, "Games where Stockfish is white", axis=ax)
    pgn.plot_move_count_histogram_cumulative(games_where_stockfish_is_black, "Games where Stockfish is black", axis=ax)

    plt.legend()
    # plt.show()
    
    pgn.plot_plycount_distribution(list_of_games)

    print(f"Time: {time.time() - start_time}")
    
    
    
    



if __name__ == "__main__":
    main()
    pass
    

    