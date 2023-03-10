from PGNGame import *
from PGNMove import *
import re
import time
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook
import textwrap

class PGNDatabase:
    '''
    Encapsulate all games parsed from a PGN file or a Excel file.
    '''

    def __init__(self, games=None):
        if games is not None:
            self.games = games
        else:
            self.games = []

    ## GETTERS ##

    def get_list_of_games(self):
        return self.games

    def get_list_of_white_wins(self):
        white_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1-0':
                white_wins.append(game)
        return white_wins

    def get_list_of_black_wins(self):
        black_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '0-1':
                black_wins.append(game)
        return black_wins

    def get_list_of_draws(self):
        draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2':
                draws.append(game)
        return draws

    def get_precentage_of_black_wins(self):
        return round((len(self.get_list_of_black_wins()) / len(self.get_list_of_games()))*100, 2)

    def get_precentage_of_white_wins(self):
        return round((len(self.get_list_of_white_wins()) / len(self.get_list_of_games()))*100, 2)

    def get_precentage_of_draws(self):
        return round((len(self.get_list_of_draws()) / len(self.get_list_of_games()))*100,2)

    def get_list_of_stockfish_wins_as_white(self):
        stockfish_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1-0' and game.lookup_meta_data('White') == 'Stockfish 15 64-bit':
                stockfish_wins.append(game)
        return stockfish_wins
    
    def get_list_of_stockfish_wins_as_black(self):
        stockfish_wins = []
        for game in self.games:

            if game.lookup_meta_data('Result') == '0-1' and game.lookup_meta_data('Black') == 'Stockfish 15 64-bit':
                stockfish_wins.append(game)
        return stockfish_wins

    def get_list_of_stockfish_losses_as_white(self):
        stockfish_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '0-1' and game.lookup_meta_data('White') == 'Stockfish 15 64-bit':
                stockfish_wins.append(game)
        return stockfish_wins

    def get_list_of_stockfish_losses_as_black(self):
        stockfish_wins = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1-0' and game.lookup_meta_data('Black') == 'Stockfish 15 64-bit':
                stockfish_wins.append(game)
        return stockfish_wins

    def get_list_of_stockfish_draws_as_white(self):
        stockfish_draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2' and game.lookup_meta_data('White') == 'Stockfish 15 64-bit':
                stockfish_draws.append(game)
        return stockfish_draws

    def get_list_of_stockfish_draws_as_black(self):
        stockfish_draws = []
        for game in self.games:
            if game.lookup_meta_data('Result') == '1/2-1/2' and game.lookup_meta_data('Black') == 'Stockfish 15 64-bit':
                stockfish_draws.append(game)
        return stockfish_draws

    def get_list_of_stockfish_wins(self):
        return self.get_list_of_stockfish_wins_as_black() + self.get_list_of_stockfish_wins_as_white()

    def get_list_of_stockfish_losses(self):
        return self.get_list_of_stockfish_losses_as_black() + self.get_list_of_stockfish_losses_as_white()
    
    def get_list_of_games_where_stockfish_is_white(self):
        stockfish_white = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('White'):
                stockfish_white.append(game)
        return stockfish_white

    def get_list_of_games_where_stockfish_is_black(self):
        stockfish_black = []
        for game in self.games:
            if "Stockfish" in game.lookup_meta_data('Black'):
                stockfish_black.append(game)
        return stockfish_black

    def get_database_of_games_where_stockfish_wins_or_draws(self):
        return PGNDatabase(self.get_list_of_stockfish_wins() + self.get_list_of_draws())

    def get_database_of_games_where_stockfish_wins(self):
        return PGNDatabase(self.get_list_of_stockfish_wins())

    def get_database_of_games_where_stockfish_losses(self):
        return PGNDatabase(self.get_list_of_stockfish_losses())

    ## GETTERS FOR OPENINGS ##

    def get_database_with_opening(self, opening):
        games_with_opening = []
        for game in self.games:
            if game.lookup_meta_data('Opening') == opening:

                games_with_opening.append(game)

        return PGNDatabase(games_with_opening)

    def get_opening_counts(self): #Returns a dictionary with keys being openings and values being the number of games with that opening
        openings = {}
        for game in self.games:
            opening = game.lookup_meta_data('Opening')
            if opening in openings:
                # If the opening is already in the dictionary, increment its value by 1
                openings[opening] += 1
            else:
                # If the opening is not in the dictionary, add it with a value of 1
                openings[opening] = 1
        return openings

    def get_openings_that_occurred_at_least_n_times(self, n):
        # Get a dictionary with keys being openings and values being the number of games with that opening
        openings = self.get_opening_counts()
        # Create a new dictionary to store the openings that occurred at least n times
        openings_that_occurred_at_least_n_times = {}
        for opening in openings:
            # If the opening occurred at least n times, add it to the new dictionary
            if openings[opening] >= n:
                # The value of the new dictionary is the number of games with that opening
                openings_that_occurred_at_least_n_times[opening] = openings[opening]

        return openings_that_occurred_at_least_n_times


    def get_standard_deviation_of_moves(self):
        amount_of_moves = []
        for game in self.games:
            amount_of_moves.append(len(game.get_moves()))
        return np.std(amount_of_moves)

    def get_mean_number_of_moves(self):
        amount_of_moves = []
        for game in self.games:
            amount_of_moves.append(len(game.get_moves()))
        return np.mean(amount_of_moves)

    def get_move_count_distribution(self, list_of_games):
        move_count_distribution = {}
        for game in list_of_games:
            move_count = len(game.get_moves())
            if move_count in move_count_distribution:
                move_count_distribution[move_count] += 1
            else:
                move_count_distribution[move_count] = 1
        return move_count_distribution

    def sort_dict(self, dict):
        return sorted(dict.items(), key=lambda x: x[0])

    def plot_move_count_histogram_cumulative(self, list_of_games, textinfo, axis=None):
        move_count_distribution = self.get_move_count_distribution(
            list_of_games)
        move_count_distribution = self.sort_dict(move_count_distribution)
        data = [(moves, games) for moves, games in move_count_distribution]
        data = sorted([(moves, games) for moves, games in data], reverse=True)
        # calculate the cumulative sum of the games for each move count
        cumulative_data = []
        cumulative_sum = 0
        for moves, games in data:
            cumulative_sum += games
            cumulative_data.append((moves, cumulative_sum))
        # plot the cumulative frequency as a function of the move count
        x = [moves for moves, games in cumulative_data]
        y = [games for moves, games in cumulative_data]
        if axis is None:  # if no axis is given, create a new one
            axis = plt.gca()  # get current axis
        axis.plot(x, y, label=textinfo)
        axis.set_xlim(15, 125)
        axis.set_xlabel('Number of Moves')
        axis.set_ylabel('Cumulative Number of Games')
        axis.set_title('Reverse Histogram of Chess Game Length')

    def plot_multiple_move_count_histogram_cumulative(self, dict_of_lists_of_games):
        for key, value in dict_of_lists_of_games.items():
            self.plot_move_count_histogram_cumulative(value, key)
        plt.legend()

    def save_histogram(self, path):
        plt.savefig(path)

    def clear_plot(self):
        plt.clf()


    ## PARSE AND COMPOSE FOR PGN AND EXCEL ##

    def compose_to_excel(self, path, sheet_name):
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = sheet_name

        i = 1
        for game in self.games:

            for key, value in game.get_meta_data().items():
                worksheet.cell(row=i, column=1, value=key)
                worksheet.cell(row=i, column=2, value=value)
                i += 1
            i += 1

            for move in game.get_moves():
                worksheet.cell(row=i, column=1, value=move.get_number())
                worksheet.cell(row=i, column=2, value=move.get_white_move())
                worksheet.cell(row=i, column=3, value=move.get_white_comment())
                worksheet.cell(row=i, column=4, value=move.get_black_move())
                worksheet.cell(row=i, column=5, value=move.get_black_comment())
                i += 1

            i += 1
        workbook.save(filename=path)

    def parse_from_excel(self, path, sheet_name):
        workbook = load_workbook(filename=path)
        worksheet = workbook[sheet_name]
        phase = 0

        meta_data = {}
        moves = []
        for row in worksheet.iter_rows(min_row=1, values_only=True):

            if phase == 0:
                if row[0] is None:
                    phase = 1
                    continue

                meta_data[row[0]] = row[1]

            if phase == 1:
                if row[0] is None:
                    phase = 2
                    continue

                number = row[0]
                white_move = row[1]
                white_move_comment = row[2]
                black_move = row[3]
                black_move_comment = row[4]

                moves.append(PGNMove(number, white_move,
                             white_move_comment, black_move, black_move_comment))

            if phase == 2:
                self.games.append(PGNGame(meta_data, moves))
                meta_data = {}
                moves = []
                phase = 0
                continue

    def compose_to_pgn(self , path):
        pgn_data = ""
        for game in self.games:
            meta_data = ""
            for key, value in game.get_meta_data().items():
                line = "[" + str(key) + " \"" + str(value) + "\"]"
                meta_data = meta_data + line + "\n"

            pgn_data += meta_data + "\n"

            moves = ""
            for move in game.get_moves():
                moves += move.get_number() + ". "

                if move.get_white_move() is not None:
                    moves += move.get_white_move() + " "

                if move.get_white_comment() is not None:
                    moves += move.get_white_comment() + " "

                if move.get_black_move() is not None:
                    moves += move.get_black_move() + " "

                if move.get_black_comment() is not None:
                    moves += move.get_black_comment() + " " 
            
            moves = textwrap.fill(moves, width=80)
            score = game.get_meta_data()["Result"]
            pgn_data += moves + score + "\n\n"
        
        pgn_file = open(path,"w")
        pgn_file.write(pgn_data)
        pgn_file.close()
    
    def parse_from_pgn(self, path):
        file = open(path, 'r')
        pgn_data = file.read()
        games = list(filter(lambda x: len(x) > 0, pgn_data.split('\n\n[')))

        game_list = []

        for game in games:
            chessgame = PGNGame({}, [])
            # Split on empty lines
            sections = list(filter(lambda x: len(x) > 0, game.split('\n\n')))
            meta_data = sections[0]
            meta_data = meta_data.split('\n')
            # Remove empty strings
            meta_data = list(filter(lambda x: len(x) > 0, meta_data))
            # Remove quatation marks
            meta_data = list(map(lambda x: x.replace(
                '[', '').replace(']', ''), meta_data))  # Remove brackets
            # Split on first space
            meta_data = list(map(lambda x: x.split(' ', 1), meta_data))
            # Remove quotes
            meta_data = list(
                map(lambda x: (x[0], x[1].replace('"', '')), meta_data))
            meta_data = dict(meta_data)
            for key, value in meta_data.items():
                chessgame.add_meta_data(key, value)

            moves = sections[1].replace('\n', ' ')

            # remove score at the end of the moves
            moves = re.sub(r'\s(1\/2-1\/2|1-0|0-1)$', '', moves)

            pattern = r'\d+\.\s[\S\s]+?(?=\d+\.\s|\Z)'  # Matches all moves
            # Define a regular expression to match the input string
            matches = re.findall(pattern, moves)
            for match in matches:
                pattern = r'(\d+)\.?\s*(\S+)\s*({.*?})?\s*(\S+)?\s*({.*?})?'
                result = re.match(pattern, match)
                if result:
                    number = result.group(1)
                    white_move = result.group(2)
                    white_move_comment = result.group(3)
                    black_move = result.group(4)
                    black_move_comment = result.group(5)
                    chessgame.add_move(PGNMove(
                        number, white_move, white_move_comment, black_move, black_move_comment))  # Add move to game
                else:
                    print("No match---------------------")
            game_list.append(chessgame)

        self.games = game_list


def main():
    time_start = time.time()

    pgn = PGNDatabase()

    pgn.parse_from_pgn("./databases/100_games.pgn")
    pgn.compose_to_pgn("./databases/100_games_composed.pgn")
    pgn.compose_to_excel("./databases/100_games_composed.xlsx", "Sheet1")
    pgn.parse_from_excel("./databases/100_games_composed.xlsx", "Sheet1")
    
    print(f"Time: {time.time() - time_start}")
    
    
if __name__ == "__main__":
    main()
    

    